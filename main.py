# -*- coding: utf-8 -*-

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
import re
from datetime import datetime
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from time import sleep
import time
import json
import os
import os.path
from modules.google_sheet import send_sheets
from tqdm import tqdm
import platform

def login_site():
    print('fazendo login...')
    login_url = 'https://admin.avec.beauty/salaojovem/admin'
    driver.get(login_url)
    # sleep(6)
    email_field = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.XPATH, '//*[@id="formEmail"]'))
    )
    # email_field = driver.find_element(By.XPATH, '//*[@id="formEmail"]')
    email_field.send_keys(credentials['email'])
    password_field = driver.find_element(By.XPATH, '//*[@id="formSenha"]')
    password_field.send_keys(credentials['password'])
    button_submit = driver.find_element(By.CSS_SELECTOR, 'button.btn')
    button_submit.click()

def get_max_option_in_select():
    try:
        select_elem = driver.find_element(By.NAME, 'tableFilter_length')
        select = Select(select_elem)
        select.select_by_value("500")
    except:
        print('error in select 500 values')

def get_infos_of_general_report_page():
    def get_report_header():
        reports_header = []
        thead = driver.find_element(By.TAG_NAME, 'thead')
        report_blocks_head = thead.find_elements(By.TAG_NAME, 'th')
        report_blocks_head.pop() # retirando ultimo elemento de click pra página
        for i in report_blocks_head:
            reports_header.append(i.text.strip())
        return reports_header

    def get_values_of_header(reports_header):
        reports = []

        tbody = driver.find_element(By.TAG_NAME, 'tbody')
        blocks_reports = tbody.find_elements(By.TAG_NAME, 'tr')
        for block in blocks_reports:
            tags_with_values = block.find_elements(By.TAG_NAME, 'td')

            code = tags_with_values[0].text.strip()
            categorie = tags_with_values[1].text.strip()
            report = tags_with_values[2].text.strip()
            description = tags_with_values[3].text.strip()
            report_links = tags_with_values[4].find_elements(By.TAG_NAME, 'a')
            link = report_links[1].get_attribute('href')
            # print(code, categorie, report, description, link)
            report = {
                reports_header[0]:code,
                reports_header[1]:categorie,
                reports_header[2]:report,
                reports_header[3]:description,
                'url':link
            }
            reports.append(report)
        
        reports_dict = {
            "reports":reports
        }

        return reports_dict
    
    sleep(2)
    reports_header = get_report_header()
    reports_dict = get_values_of_header(reports_header)
    return reports_dict

def filter_lists(reports):
    print('filtrando categorias...')
    categories = []
    lists = {}
    for r in reports:
        if r['Categoria'] not in categories:
            categories.append(r['Categoria'])
    for c in categories:
        lists[c] = []
        for report in reports:
            if report['Categoria'] == c:
                lists[c].append(report)
    make_json('reports_lists.json', lists)
    return lists

def table_to_dict(headers, values):
    data = {}
    for header, value in zip(headers, values):
        data[header] = value
    return data

def set_only_values_between(last_date_updated=None):
    if last_date_updated is None:
        last_date_updated = '01/01/2020'
    atual_date = datetime.now().strftime("%d/%m/%Y")
    
    sleep(3)
    try:
        field_initial_date = driver.find_element(By.NAME, 'inicio')
        field_final_date = driver.find_element(By.NAME, 'fim')
        search_button = driver.find_element(By.CSS_SELECTOR, '.btn-info')
    except:
        return None

    print(f"Data inicial: {last_date_updated}")
    print(f"Data final: {atual_date}")

    driver.execute_script("arguments[0].value = arguments[1];", field_initial_date, last_date_updated)
    driver.execute_script("arguments[0].value = arguments[1];", field_final_date, atual_date)

    # Validando se os valores foram inseridos corretamente
    initial_date_value = field_initial_date.get_attribute('value')
    final_date_value = field_final_date.get_attribute('value')

    print(f"Valor no campo inicial: {initial_date_value}")
    print(f"Valor no campo final: {final_date_value}")

    if initial_date_value == last_date_updated and final_date_value == atual_date:
        search_button.click()
        sleep(12)
    else:
        print("Os valores dos campos não foram atualizados corretamente.")

def get_infos_in_report_page(link_report_page, relatorie, last_date_updated=None):

    def get_headers():
        head_titles = []
        div_thead = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '#tableFilter > thead:nth-child(1) > tr:nth-child(1)'))
        )
        div_thead = driver.find_element(By.CSS_SELECTOR, '#tableFilter > thead:nth-child(1) > tr:nth-child(1)')
        tags_head = div_thead.find_elements(By.TAG_NAME, 'th')
        for tag in tags_head:
            title_head = tag.find_element(By.TAG_NAME, 'div').text
            head_titles.append(title_head)
        return head_titles
        
    def tables_to_dict(headers, values):
        data_list = []
        num_headers = len(headers)
        num_values = len(values)
        group_size = num_values // num_headers
        
        for i in range(0, num_values, num_headers):
            group_values = values[i:i+num_headers]
            data = {}
            for header, value in zip(headers, group_values):
                data[header] = value
            data_list.append(data)
        
        return data_list 

    def get_value_of_headers(headers, relatorie):
        values = []
        # sleep(12)
        tbody = driver.find_element(By.CSS_SELECTOR, '#tableFilter > tbody:nth-child(2)')
        tags_rows = tbody.find_elements(By.TAG_NAME, 'td')
        if len(tags_rows) == 1 and tags_rows[0].get_attribute('class') == 'dataTables_empty':
            print('pulado')
            return None
        while True:
            tbody = driver.find_element(By.CSS_SELECTOR, '#tableFilter > tbody:nth-child(2)')
            tags_rows = tbody.find_elements(By.TAG_NAME, 'td')
            # sleep(2)
            # if len(tags_rows) == 1 and tags_rows[0].get_attribute('class') == 'dataTables_empty':
            #     print('pulado')
            #     return None
            for tag in tags_rows:
                values.append(tag.text.strip())

            next_tab = driver.find_element(By.CSS_SELECTOR, 'li.next')
            # print(next_tab.text)
            try:
                class_name = next_tab.get_attribute('class')

                if 'disabled' in str(class_name):
                    # print('sem mais páginas pra pegar')
                    break
                else:
                    a_tag = next_tab.find_element(By.TAG_NAME, 'a')
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    #sleep(1)
                    # print('conseguiu pegar o link do proximo')
                    try:
                        a_tag.click()
                    except:
                        # print('erro ao clicar no proximo elemento')
                        break
            except NoSuchElementException:
                print('proximo elemento não encontrado')
                break
        # print(f'values: {len(values)}')
        print(f'número de valores encontrados: {int(len(values)/len(headers))}')
        list_result = tables_to_dict(headers, values)
        dict_result = {
            relatorie:list_result
        }

        return dict_result
    
    # test com maior quantidade de dados possivel
    # link_report_page='https://admin.avec.beauty/admin/relatorio/0125'
    # login_site()
    driver.get(link_report_page)
    print(link_report_page)
    if last_date_updated is not None:
        set_only_values_between(last_date_updated)
    else:
        set_only_values_between()
    get_max_option_in_select()
    headers = get_headers()
    report_dict = get_value_of_headers(headers, relatorie)

    return report_dict

def get_reports(last_date_updated=None):
    def get_lists_json():
        print('buscando e ordenando links...')
        reports_section = WebDriverWait(driver, 40).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="itemMenuRelatorios"]'))
        )
        reports_section.click()
        
        # Re-localizando o elemento para evitar o erro stale element reference
        all_reports_section = WebDriverWait(driver, 40).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="relatorios"]'))
        )
        all_reports_section.click()
        
        get_max_option_in_select()
        reports_dict = get_infos_of_general_report_page()
        reports = reports_dict['reports']
        reports_lists = filter_lists(reports)
        return reports_lists
    reports_lists = get_lists_json()
    make_json('reports_lists.json',reports_lists)
    
    # temporario 
    # reports_lists = load_json('reports_lists.json')

    reports = []

    if last_date_updated is not None:
        print(f'Buscando relatórios de {last_date_updated} até {datetime.now().strftime("%d/%m/%Y")}')

    def get_infos(reports_lists):
        # buscar apenas informações de relatórios de uma categoria especifica
        # priority_category_list = ['Clientes', 'Financeiro']

        def push_report(report):
            # if report['Categoria'] not in priority_category_list:
            #     return None
            report_dict = get_infos_in_report_page(report['url'], report['Relatório'], last_date_updated)
            if report_dict is not None:
                return report_dict
            else:
                return None
        
       
        for j, (category, reports_list) in enumerate(reports_lists.items()):
            category_reports = []
            # if category not in priority_category_list:
            #     continue
            for i, report in enumerate(reports_list):
                report_dict = push_report(report)
                if report_dict is not None:
                    category_reports.append(report_dict)
            if category_reports:
                reports_by_category = {
                    category: category_reports
                }
                reports.append(reports_by_category)

    get_infos(reports_lists)
    make_json('reports.json', {'reports':reports})

    # temporario
    # reports = load_json('reports.json')
    # return reports

    return {'reports':reports}

def make_json(file_name, dict):
    with open(file_name, 'w', encoding="utf8") as file:
        json.dump(dict, file, indent=4, ensure_ascii=False)

def load_json(file_name):
    with open(file_name, 'r', encoding="utf8") as file:
        return json.load(file)

def find_variable_name(value):
    for name, val in globals().items():
        if val is value:
            return name
    return None

def make_excel_table(reports:dict):
    folder_name = './planilhas'

    def make_folder():
        print('criando pasta de planilhas...')
        if os.path.isdir('./planilhas'):
            pass
        else:
            os.mkdir('./planilhas')
    
    def values_already_exists(book_name, sheet_name, values_to_insert):
        print('verificando existência de valores nas planilhas...')
        book = openpyxl.load_workbook(book_name)
        sheet = book[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            if list(row) == values_to_insert:
                    print('valor já existe, não adicionado')
                    return True
            else:
                continue

    import openpyxl

    def update_value_in_excel(excel_file):
        # Abrir o arquivo Excel
        workbook = openpyxl.load_workbook(excel_file)
        
        # Iterar por todas as planilhas
        for sheet in workbook.worksheets:
            # Listar para armazenar as linhas modificadas
            modified_rows = []
            
            # Iterar por todas as linhas e células
            for row in sheet.iter_rows(values_only=True):
                new_row = []
                for cell in row:
                    if isinstance(cell, str):
                        if cell.count('"') == 1:
                            new_value = cell.replace('"', '')
                        elif cell.endswith('.0') and cell.count('.0') == 1:
                            new_value = int(cell.replace('.0', ''))
                        elif cell.isdigit():
                            new_value = int(cell)
                        else:
                            new_value = cell
                    else:
                        new_value = cell
                    new_row.append(new_value)
                modified_rows.append(new_row)
            
            # Atualizar as células da planilha com os novos valores
            for i, row in enumerate(modified_rows, start=1):
                for j, value in enumerate(row, start=1):
                    sheet.cell(row=i, column=j, value=value)
        
        # Salvar o arquivo Excel modificado
        workbook.save(excel_file)

    def make_excel(reports, folder_name):
        print('criando planilhas...')
        reports = reports['reports']
        for categories_items in reports:
            for category_name, values in categories_items.items():
                book = openpyxl.Workbook()
                for report in values:
                    for key, value in report.items():
                        title = re.sub(INVALID_TITLE_REGEX, '_', key)
                        if title != 'category':
                            book.create_sheet(title)
                        categorie_page = book[title]
                        headers_added = False
                        for item in value:
                            headers = []
                            tuples = []
                            if type(item) != dict:
                                print(item)
                                continue
                            for key_, value_ in item.items():
                                headers.append(key_)
                                tuples.append(value_) 
                            if headers_added is False:
                                categorie_page.append(headers)
                                headers_added = True
                            categorie_page.append(tuples)
                pattern_sheet = book['Sheet']
                book.remove(pattern_sheet)
                book.save(f'{folder_name}/{category_name}.xlsx')
                update_value_in_excel(f'{folder_name}/{category_name}.xlsx')
                print(f'planilha: {category_name}.xlsx criada')

    make_folder()
    make_excel(reports, folder_name)
    send_sheets()
    
if __name__ == '__main__':
    start_time = time.time()
    with open('settings.json', 'r') as file:
        credentials = json.load(file).get('credentials')

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('log-level=3')
    if platform.system() == 'Linux':
        chrome_driver_path = Service('./chromedriver')
        driver = webdriver.Chrome(service=chrome_driver_path, options=chrome_options)
    else:
        chrome_driver_path = Service('./chromedriver_arm')
        driver = webdriver.Chrome(service=chrome_driver_path, options=chrome_options)
    wait = WebDriverWait(driver, 5)
    if os.path.isfile('./settings.json'):
        settings = load_json('./settings.json')
        login_site()
        reports = get_reports(settings['last_date_updated'])
    else:
        login_site()
        reports = get_reports()
    JSON = load_json('settings.json')
    print(JSON)
    JSON['last_date_updated'] = datetime.now().strftime("%d/%m/%Y")
    make_json('settings.json', JSON)
    end_time = time.time()
    duration = (end_time - start_time) / 60
    print("Tempo decorrido: ", duration, "minutos")
    # with open('reports.json', 'r') as file:
        # reports = json.load(file)
    # make_excel_table(reports)